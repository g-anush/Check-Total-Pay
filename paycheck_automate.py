import csv
from openpyxl import load_workbook
from datetime import datetime, timedelta


def calculate_overtime_pay(hourly_rate, total_overtime_hours):
    if total_overtime_hours <= 50:
        return total_overtime_hours * (hourly_rate + 0.20 * hourly_rate)
    else:
        return ((50 * (hourly_rate * 0.20 + hourly_rate)) + ((total_overtime_hours - 50) * (hourly_rate * 0.15 + hourly_rate)))


def calculate_night_shift_pay(hourly_rate, total_night_shift_overtime_hours):
    return (total_night_shift_overtime_hours * (hourly_rate * 0.25 + hourly_rate))


def calculate_holiday_pay(hourly_rate, total_holiday_overtime_hours):
    return (total_holiday_overtime_hours * (hourly_rate * 0.30 + hourly_rate))

# To check if in a time interval night shift lies or not


def is_night_shift(time1, time2):
    return (
        (time1.hour <= 2 and time2.hour >= 5) or
        (time1.hour <= 2 and time2.hour <= 5) or
        (time1.hour >= 2 and time2.hour <= 5) or
        ((5 > time1.hour >= 2) and time2.hour >= 5)
    )


def calculate_weekly_salary(entries):
    week_day = 0                                # Will store day number of the week
    week_flag = False                           # Indicates whether the employee has completed 5 days of work in a week
    hourly_rate = 100                           # Rupees per hour
    regular_hours_per_day = 9                   # Minimum hours for overtime eligibility in a day
    overtime_threshold = 45                     # Minimum hours for overtime eligibility in a week
    hours_of_daily_work_remaining = 0           # Remaining hour to complete weekly work
    not_present = 0                             # Number of days employee has taken leave
    daily_hours_worked = 0                      # Hours worked in a day
    daily_ordinary_hours_worked = 0             # Ordinary hours worked in a day
    daily_overtime_hours = 0                    # Overtime hours in a day
    ordinary_breaks_taken = 0
    overtime_breaks_taken = 0
    daily_night_shift_overtime_hours = 0        # Overtime hours for night shift in a day
    daily_holiday_overtime_hours = 0            # Overtime hours for a holiday
    total_salary = 0                            # Total salary for the week
    total_hours_worked = 0                      # Total hours worked in a week
    total_ordinary_hours_worked = 0           # Total ordinary hours worked in a week
    total_overtime_hours = 0                    # Total overtime hours in a week
    total_night_shift_overtime_hours = 0        # Total night shift hours in a week
    total_holiday_overtime_hours = 0            # Total holiday hours in a week
    week = int(entries[0]["Week"])              # Week number

    for entry in entries:
        if int(entry["Week"]) == week:
            week_day += 1
        else:
            week = int(entry["Week"])
            week_day = 1

        print("**********")
        print(f"Day {week_day} of Week {week}")

        # Skip entries having NaN values
        if entry["Time_In"] == "NaN" or entry["Time_Out"] == "NaN":
            not_present += 1
            print("Employee has taken leave for this day.")
            continue

        time_in_str = entry["Time_In"].strftime("%I:%M %p")
        time_out_str = entry["Time_Out"].strftime("%I:%M %p")
        national_holiday = (entry["National_Holiday"].upper() == "YES")

        time_in = entry["Date"].strftime("%d-%b-%y") + " " + time_in_str
        time_out = entry["Date"].strftime("%d-%b-%y") + " " + time_out_str
        night_shift_start = entry["Date"].strftime(
            "%d-%b-%y") + " " + '02:00 AM'
        night_shift_end = entry["Date"].strftime("%d-%b-%y") + " " + '05:00 AM'

        time_in = datetime.strptime(time_in, "%d-%b-%y %I:%M %p")
        time_out = datetime.strptime(time_out, "%d-%b-%y %I:%M %p")
        night_shift_start = datetime.strptime(
            night_shift_start, "%d-%b-%y %I:%M %p")
        night_shift_end = datetime.strptime(
            night_shift_end, "%d-%b-%y %I:%M %p")

        if time_out < time_in:
            time_out += timedelta(days=1)

        # Calculate total hours worked, excluding breaks in a day
        daily_hours_worked = (time_out - time_in).total_seconds() / 3600

        # Calculate ordinary working hours in a day
        daily_ordinary_hours_worked = min(daily_hours_worked, 9)

        ordinary_breaks_taken = (daily_ordinary_hours_worked // 3.5)
        # Deduct break time of 30 minutes after every 3.5 hours
        daily_ordinary_hours_worked -= ordinary_breaks_taken * 0.5

        # Calculate daily overtime hours
        if daily_hours_worked > 9:
            daily_overtime_hours = daily_hours_worked - daily_ordinary_hours_worked
            # Deduct of break time as this time will be added to ordinary  hours worked
            daily_overtime_hours -= ordinary_breaks_taken * 0.5
            daily_ordinary_hours_worked += ordinary_breaks_taken * 0.5
            if daily_overtime_hours < 0:
                hours_of_daily_work_remaining += abs(daily_overtime_hours)
            # Until minimum daily hours of per day work requirement is not fulfilled
            # overtime hours will not be counted
            if hours_of_daily_work_remaining > 0:
                daily_overtime_hours -= hours_of_daily_work_remaining
                hours_of_daily_work_remaining = 0
                if daily_overtime_hours < 0:
                    hours_of_daily_work_remaining = abs(daily_overtime_hours)
                    daily_overtime_hours = 0
            overtime_breaks_taken = daily_overtime_hours//3.5
            daily_overtime_hours -= overtime_breaks_taken * 0.5
        else:
            hours_of_daily_work_remaining = 9 - daily_ordinary_hours_worked

        print(f"Daily hours worked: {daily_hours_worked}")
        print(f"Daily ordinary hours worked: {daily_ordinary_hours_worked}")
        print(f"Breaks taken: {ordinary_breaks_taken + overtime_breaks_taken}")
        print(f"Total hours worked: {daily_hours_worked}")
        print(f"Daily overtime hours: {daily_overtime_hours}")
        print(
            f"Hours of daily work remaining: {hours_of_daily_work_remaining}")

        # Update overtime, night shift, and holiday hours
        if (week_day - not_present) > 5:
            # week_flag indicates the employee has completed 5 days of work in a week
            week_flag = True

        if (not week_flag and (2 <= time_in.hour <= 5 or 2 <= time_out.hour <= 5) and (daily_hours_worked >= 9 and total_hours_worked >= week_day*9)):
            if ((time_out.hour >= 5 and time_in.hour <= 2)):
                daily_night_shift_overtime_hours += (5-2)
            elif time_out.hour > 5 and time_in.hour >= 2:
                daily_night_shift_overtime_hours += (
                    night_shift_end - time_in).total_seconds() / 3600
            elif time_out.hour <= 5 and time_in.hour <= 2:
                daily_night_shift_overtime_hours += (
                    time_out - night_shift_start).total_seconds() / 3600
            else:
                daily_night_shift_overtime_hours += (
                    time_out - time_in).total_seconds() / 3600
        elif week_flag and total_hours_worked >= overtime_threshold:
            if ((time_out.hour >= 5 and time_in.hour <= 2)):
                daily_night_shift_overtime_hours += (5-2)
            elif time_out.hour > 5 and time_in.hour >= 2:
                daily_night_shift_overtime_hours += (
                    night_shift_end - time_in).total_seconds() / 3600
            elif time_out.hour <= 5 and time_in.hour <= 2:
                daily_night_shift_overtime_hours += (
                    time_out - night_shift_start).total_seconds() / 3600
            else:
                daily_night_shift_overtime_hours += (
                    time_out - time_in).total_seconds() / 3600

        breaks_taken = daily_night_shift_overtime_hours//3.5
        daily_night_shift_overtime_hours -= (breaks_taken * 0.5)

        if national_holiday:
            daily_holiday_overtime_hours = daily_overtime_hours

        total_hours_worked += daily_hours_worked
        total_ordinary_hours_worked += daily_ordinary_hours_worked
        if daily_overtime_hours > 0:
            total_overtime_hours += daily_overtime_hours
        total_night_shift_overtime_hours += daily_night_shift_overtime_hours
        total_holiday_overtime_hours += daily_holiday_overtime_hours
        daily_total_income, daily_overtime_income = 0, 0

        print(
            f"Total ordinary hours payment for the day: {daily_ordinary_hours_worked * hourly_rate}")

        if daily_overtime_hours > 0:
            daily_overtime_income = calculate_overtime_pay(
                hourly_rate, daily_overtime_hours)
            print(
                f"Total overtime hours payment for the day: {calculate_overtime_pay(hourly_rate, daily_overtime_hours)}")
        else:
            print(
                f"Total overtime hours payment for the day: 0.00")

        print(
            f"Total night shift hours payment for the day: {calculate_night_shift_pay(hourly_rate, daily_night_shift_overtime_hours)}")
        print(
            f"Total holiday hours payment for the day: {calculate_holiday_pay(hourly_rate, daily_holiday_overtime_hours)}")

        daily_ordinary_income = daily_ordinary_hours_worked * hourly_rate
        daily_night_shift_income = calculate_night_shift_pay(
            hourly_rate, daily_night_shift_overtime_hours)
        daily_holiday_income = calculate_holiday_pay(
            hourly_rate, daily_holiday_overtime_hours)
        daily_income = daily_ordinary_income + daily_overtime_income + \
            daily_night_shift_income + daily_holiday_income

        print(f"Daily Income: {daily_income}")

        daily_hours_worked = 0
        daily_ordinary_hours_worked = 0
        daily_overtime_hours = 0
        daily_night_shift_overtime_hours = 0
        daily_holiday_overtime_hours = 0

        print("**********")

    # Calculate salary for the week
    total_salary += total_ordinary_hours_worked * hourly_rate
    total_salary += calculate_overtime_pay(hourly_rate, total_overtime_hours)
    total_salary += calculate_night_shift_pay(
        hourly_rate, total_night_shift_overtime_hours)
    total_salary += calculate_holiday_pay(hourly_rate,
                                          total_holiday_overtime_hours)

    print(f"Total salary: {total_salary}")

    return total_salary


# Read time sheet from CSV file
time_sheet = []
excel_file = "time_sheet.xlsx"
workbook = load_workbook(excel_file, read_only=True)
sheet = workbook["Sheet1"]
for row in sheet.iter_rows(min_row=2, values_only=True):
    entry = {
        "Week": row[0],
        "Date": row[1],
        "Time_In": row[2],
        "Time_Out": row[3],
        "National_Holiday": row[4]
    }
    time_sheet.append(entry)

# Group entries by week
entries_by_week = {}
for entry in time_sheet:
    week = int(entry["Week"])
    if week not in entries_by_week:
        entries_by_week[week] = []
    entries_by_week[week].append(entry)

# Calculate and print the total salary for each week
biweekly_salary, grand_total_salary, flag = 0, 0, 0
for week, entries in entries_by_week.items():
    weekly_salary = calculate_weekly_salary(entries)
    biweekly_salary += weekly_salary
    flag += 1
    if flag == 2:
        print(f"Biweekly Salary: Rupees {biweekly_salary:.2f}")
        biweekly_salary, flag = 0, 0
    grand_total_salary += weekly_salary
    print(f"Week {week} Salary: Rupees {weekly_salary:.2f}")

# Print the grand total salary
print("********************")
print(f"Grand Total Salary: Rupees {grand_total_salary:.2f}")
print("********************")
